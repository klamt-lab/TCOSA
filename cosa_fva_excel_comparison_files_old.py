import copy
import math
import openpyxl
from helper import json_load
from cosa_load_model_data import load_model_data
from cosa_get_model_with_nadx_scenario import cosa_get_model_with_nadx_scenario


core_ids = [
    "EX_glc__D_e",
    "GLCptspp",
    "GLCt2pp",
    "HEX1",
    "G6PDH2r",
    "PGL",
    "F6PA",
    "PGI",
    "GND",
    "PFK",
    "FBP",
    "RPE",
    "RPI",
    "GLYCDx",
    "FBA",
    "TKT2",
    "TKT1",
    "TALA",
    "TPI",
    "G3PD2",
    "G3PD5",
    "GLYK",
    "EX_glyc_e",
    "PGK",
    "GAPD",
    "PGM",
    "ENO",
    "EDA",
    "EDD",
    "EX_h_e",
    "EX_h2o_e",
    "EX_co2_e",
    "PPC",
    "PPCK",
    "PYK",
    "PPS",
    "MGSA",
    "GLYOX3",
    "LDH_D",
    "FHL",
    "PFL",
    "ME2",
    "ME1",
    "PDH",
    "POX",
    "PTAr",
    "ACALD",
    "ACKr",
    "ALCD2x",
    "EX_ac_e",
    "EX_etoh_e",
    "CS",
    "EX_succ_e",
    "SUCCt2_2pp",
    "SUCDi",
    "FRD2",
    "FUM",
    "MDH",
    "MALS",
    "SUCOAS",
    "AKGDH",
    "ICL",
    "ACONTa",
    "ACONTb",
    "ICDHyr",
    "ADK1",
    "ATPM",
    "ATPS4rpp",
    "EX_o2_e",
    "CYTBO3_4pp",
    "NADH16pp",
    "NADH17pp",
    "NADTRHD",
    "THD2pp",
    "EX_lac__D_e",
    "EX_h2_e",
    "EX_for_e",
    "SUCCt2_3pp",
    "SUCCt1pp",
    "BIOMASS_Ec_iML1515_core_75p37M"
]


table_headers = [
    "variability__aerobic_TEST_0_818_PAPERCONCS_",
    "variability__aerobic_TEST_0_818_STANDARDCONCS_",
    "variability__anaerobic_TEST_0_321_STANDARDCONCS_",
    "variability__anaerobic_TEST_0_321_PAPERCONCS_",
]

color_dark_red = "8B0000"
fill_dark_red = openpyxl.styles.PatternFill(start_color=color_dark_red, end_color=color_dark_red, fill_type="solid")
color_light_red = "FF7F7F"
fill_light_red = openpyxl.styles.PatternFill(start_color=color_light_red, end_color=color_light_red, fill_type="solid")
color_green = "00FF00"
fill_green = openpyxl.styles.PatternFill(start_color=color_green, end_color=color_green, fill_type="solid")
color_light_blue = "ADD8E6"
fill_light_blue = openpyxl.styles.PatternFill(start_color=color_light_blue, end_color=color_light_blue, fill_type="solid")
color_dark_blue = "394D6D"
fill_dark_blue = openpyxl.styles.PatternFill(start_color=color_dark_blue, end_color=color_dark_blue, fill_type="solid")
color_white = "FFFFFF"
fill_white = openpyxl.styles.PatternFill(start_color=color_white, end_color=color_white, fill_type="solid")
color_black = "000000"
fill_black = openpyxl.styles.PatternFill(start_color=color_black, end_color=color_black, fill_type="solid")

italic = openpyxl.styles.Font(italic=True)
bold = openpyxl.styles.Font(bold=True)
border = openpyxl.styles.Border(left=openpyxl.styles.Side(border_style='thin', color='000000'))

all_base_ids, cobra_model_aerobic, concentration_values_free, concentration_values_paper,\
standardconc_dG0_values, paperconc_dG0_values,\
num_nad_and_nadp_reactions, num_nad_base_ids, num_nadp_base_ids,\
ratio_constraint_data, nad_base_ids, nadp_base_ids, used_growth, zeroed_reaction_ids = load_model_data(anaerobic=False, expanded=False)

all_base_ids, cobra_model_anaerobic, concentration_values_free, concentration_values_paper,\
standardconc_dG0_values, paperconc_dG0_values,\
num_nad_and_nadp_reactions, num_nad_base_ids, num_nadp_base_ids,\
ratio_constraint_data, nad_base_ids, nadp_base_ids, used_growth, zeroed_reaction_ids = load_model_data(anaerobic=True, expanded=False)


for table_header in table_headers:
    if "aerobic" in table_header:
        original_cobra_model = copy.deepcopy(cobra_model_aerobic)
    else:
        original_cobra_model = copy.deepcopy(cobra_model_anaerobic)

    cobra_model_1 = copy.deepcopy(original_cobra_model)
    cobra_model_2 = copy.deepcopy(original_cobra_model)
    cobra_model_3 = copy.deepcopy(original_cobra_model)

    condition_models = {
        "Wild-type": cosa_get_model_with_nadx_scenario(nadx_scenario="WILDTYPE", cobra_model=cobra_model_1),
        "Single-cofactor": cosa_get_model_with_nadx_scenario(nadx_scenario="SINGLE_COFACTOR", cobra_model=cobra_model_2),
        "Flexible": cosa_get_model_with_nadx_scenario(nadx_scenario="FLEXIBLE", cobra_model=cobra_model_3),
    }

    json_tuples = [
        (
            "Wild-type",
            json_load("./cosa/"+table_header+"WILDTYPE.json"),
            json_load("./cosa/"+table_header.replace("__", "_STOICHIOMETRIC__").replace("_STANDARDCONCS", "").replace("_PAPERCONCS", "")+"WILDTYPE.json")
        ),
        (
            "Single-cofactor",
            json_load("./cosa/"+table_header+"SINGLE_COFACTOR.json"),
            json_load("./cosa/"+table_header.replace("__", "_STOICHIOMETRIC__").replace("_STANDARDCONCS", "").replace("_PAPERCONCS", "")+"SINGLE_COFACTOR.json")
        ),
        (
            "Flexible",
            json_load("./cosa/"+table_header+"FLEXIBLE.json"),
            json_load("./cosa/"+table_header.replace("__", "_STOICHIOMETRIC__").replace("_STANDARDCONCS", "").replace("_PAPERCONCS", "")+"FLEXIBLE.json")
        ),
    ]

    json_metadata = {}
    for json_tuple in json_tuples:
        condition = json_tuple[0]
        data = json_tuple[1]
        for target in ("OptMDF", "OptSubMDF"):
            if target not in json_metadata.keys():
                json_metadata[target] = {}
            for elements in data[target]:
                var_id = elements[0]
                min_value = elements[1]
                max_value = elements[2]
                if var_id not in json_metadata[target].keys():
                    json_metadata[target][var_id] = {}

                json_metadata[target][var_id][condition] = {}
                json_metadata[target][var_id][condition]["min"] = min_value
                json_metadata[target][var_id][condition]["max"] = max_value
        data_stoich = json_tuple[2]
        for target in ("OptMDF", "OptSubMDF"):
            for elements in data_stoich["stoichiometric"]:
                var_id = elements[0]
                min_value = elements[1]
                max_value = elements[2]

                json_metadata[target][var_id][condition]["min_stoich"] = min_value
                json_metadata[target][var_id][condition]["max_stoich"] = max_value

    wb = openpyxl.Workbook()
    for target in ("OptMDF", "OptSubMDF"):
        ws = wb.create_sheet(target)

        targetdata = json_metadata[target]
        varnames = list(targetdata.keys())
        f_vars = [x for x in varnames if x.startswith("f_var_")]
        f_vars_core = []
        for f_var in f_vars:
            reac_var = f_var.replace("f_var_", "")
            for core_id in core_ids:
                if reac_var.startswith(core_id):
                    f_vars_core.append(f_var)
                    break
        f_vars_noncore = [x for x in f_vars if x not in f_vars_core]
        f_vars_core.sort()
        f_vars_noncore.sort()
        f_vars = f_vars_core + f_vars_noncore


        cell = ws.cell(3, 1)
        cell.value = "Reaction ID"
        cell.font = italic

        cell = ws.cell(3, 2)
        cell.value = "Reaction string"
        cell.font = italic

        tested_MDFs = {
            condition: round(targetdata["var_B" if target=="OptMDF" else "var_B2"][condition]["max"], 4)
            for condition in targetdata["var_B"].keys()
        }

        headers = list(targetdata[varnames[0]].keys())
        headers_to_column = {}
        n_column = 3
        for header in headers:
            headers_to_column[header] = n_column
            cell = ws.cell(1, n_column)
            cell.value = header
            cell.font = bold
            cell.border = border
            cell = ws.cell(3, n_column)
            cell.value = "min FVA"
            cell.border = border
            cell.font = italic
            cell = ws.cell(3, n_column+1)
            cell.value = "max FVA"
            cell.font = italic
            cell = ws.cell(3, n_column+2)
            cell.value = "min TFVA"
            cell.font = italic
            cell = ws.cell(3, n_column+3)
            cell.value = "max TFVA"
            cell.font = italic
            cell = ws.cell(3, n_column+4)
            cell.value = "max df"
            cell.font = italic

            cell = ws.cell(2, n_column)
            cell.value = "MDF:"
            if target == "OptMDF":
                cell.font = bold
            cell.border = border
            cell = ws.cell(2, n_column+1)
            cell.value = targetdata["var_B"][header]["max"]

            cell = ws.cell(2, n_column+2)
            cell.value = "SubMDF:"
            if target == "OptSubMDF":
                cell.font = bold
            cell = ws.cell(2, n_column+3)
            cell.value = targetdata["var_B2"][header]["max"]

            mdf = tested_MDFs[header]

            n_column += 5

        max_width = 5
        max_width_str = 5
        current_line = 4


        for f_var in f_vars:
            reac_id = f_var.replace("f_var_", "")
            max_width = max(len(reac_id)*1.2, max_width)
            ws.column_dimensions['A'].width = max_width

            cell = ws.cell(current_line, 1)
            cell.value = reac_id
            cell.alignment = openpyxl.styles.Alignment(horizontal='right')

            cell = ws.cell(current_line, 2)
            cell.value = condition_models[condition].reactions.get_by_id(reac_id).reaction
            max_width_str = max(len(str(condition_models[condition].reactions.get_by_id(reac_id).reaction))*0.5, max_width_str)
            cell.alignment = openpyxl.styles.Alignment(horizontal='right')
            ws.column_dimensions['B'].width = max_width_str

            for condition in targetdata[f_var].keys():
                model_min_flux = condition_models[condition].reactions.get_by_id(reac_id).lower_bound
                model_max_flux = condition_models[condition].reactions.get_by_id(reac_id).upper_bound

                min_fva_flux_value = targetdata[reac_id][condition]["min_stoich"]
                max_fva_flux_value = targetdata[reac_id][condition]["max_stoich"]

                min_tfva_flux_value = targetdata[reac_id][condition]["min"]
                max_tfva_flux_value = targetdata[reac_id][condition]["max"]

                min_df_value = targetdata[f_var][condition]["min"]
                max_df_value = targetdata[f_var][condition]["max"]

                # Is blocked anyway: Black
                if (model_min_flux == 0.0) and (model_max_flux == 0.0):
                    cell_filler = fill_black
                # Is essential in FVA? -> Dark blue
                elif min_fva_flux_value > 1e-8:
                    cell_filler = fill_dark_blue
                # Is essential in TFVA only? -> Light blue
                elif min_tfva_flux_value > 1e-8:
                    cell_filler = fill_light_blue
                # Is blocked?
                elif max_tfva_flux_value < 1e-8:
                    # Blocked in FVA too? -> Dark red
                    if max_fva_flux_value < 1e-8:
                        cell_filler = fill_dark_red
                    # Blocked in TFVA only? -> Light red
                    else:
                        cell_filler = fill_light_red
                # Is active?
                else:
                    cell_filler = fill_green

                cell = ws.cell(current_line, headers_to_column[condition])
                cell.value = round(min_fva_flux_value, 8)
                cell.fill = cell_filler
                cell.border = border
                #
                cell = ws.cell(current_line, headers_to_column[condition]+1)
                cell.value = round(max_fva_flux_value, 8)
                cell.fill = cell_filler

                cell = ws.cell(current_line, headers_to_column[condition]+2)
                cell.value = round(min_tfva_flux_value, 8)
                cell.fill = cell_filler
                #
                cell = ws.cell(current_line, headers_to_column[condition]+3)
                cell.value = round(max_tfva_flux_value, 8)
                cell.fill = cell_filler

                cell = ws.cell(current_line, headers_to_column[condition]+4)
                cell.value = round(max_df_value, 8)
                cell.fill = cell_filler

                if (round(max_df_value, 4) == tested_MDFs[condition]):
                    cell.font = bold

                    cell = ws.cell(current_line, 1)
                    cell.font = bold

            current_line += 1
        ws.freeze_panes = "A4"



    del(wb["Sheet"])
    wb.save(table_header+".xlsx")
