import copy
import math
import openpyxl
from helper import json_load
from cosa_load_model_data import load_model_data
from cosa_get_model_with_nadx_scenario import cosa_get_model_with_nadx_scenario


core_ids = [
    "EX_glc__D_e",
    "GLCptspp",
    "GLCt2pp",
    "HEX1",
    "G6PDH2r",
    "PGL",
    "F6PA",
    "PGI",
    "GND",
    "PFK",
    "FBP",
    "RPE",
    "RPI",
    "GLYCDx",
    "FBA",
    "TKT2",
    "TKT1",
    "TALA",
    "TPI",
    "G3PD2",
    "G3PD5",
    "GLYK",
    "EX_glyc_e",
    "PGK",
    "GAPD",
    "PGM",
    "ENO",
    "EDA",
    "EDD",
    "EX_h_e",
    "EX_h2o_e",
    "EX_co2_e",
    "PPC",
    "PPCK",
    "PYK",
    "PPS",
    "MGSA",
    "GLYOX3",
    "LDH_D",
    "FHL",
    "PFL",
    "ME2",
    "ME1",
    "PDH",
    "POX",
    "PTAr",
    "ACALD",
    "ACKr",
    "ALCD2x",
    "EX_ac_e",
    "EX_etoh_e",
    "CS",
    "EX_succ_e",
    "SUCCt2_2pp",
    "SUCDi",
    "FRD2",
    "FUM",
    "MDH",
    "MALS",
    "SUCOAS",
    "AKGDH",
    "ICL",
    "ACONTa",
    "ACONTb",
    "ICDHyr",
    "ADK1",
    "ATPM",
    "ATPS4rpp",
    "EX_o2_e",
    "CYTBO3_4pp",
    "NADH16pp",
    "NADH17pp",
    "NADTRHD",
    "THD2pp",
    "EX_lac__D_e",
    "EX_h2_e",
    "EX_for_e",
    "SUCCt2_3pp",
    "SUCCt1pp",
    "BIOMASS_Ec_iML1515_core_75p37M"
]


table_headers = [
    "variability__aerobic_TEST_0_818_PAPERCONCS_",
    "variability__aerobic_TEST_0_818_STANDARDCONCS_",
    "variability__anaerobic_TEST_0_321_STANDARDCONCS_",
    "variability__anaerobic_TEST_0_321_PAPERCONCS_",
]

color_dark_red = "8B0000"
fill_dark_red = openpyxl.styles.PatternFill(start_color=color_dark_red, end_color=color_dark_red, fill_type="solid")
color_light_red = "FF7F7F"
fill_light_red = openpyxl.styles.PatternFill(start_color=color_light_red, end_color=color_light_red, fill_type="solid")
color_green = "00FF00"
fill_green = openpyxl.styles.PatternFill(start_color=color_green, end_color=color_green, fill_type="solid")
color_light_blue = "ADD8E6"
fill_light_blue = openpyxl.styles.PatternFill(start_color=color_light_blue, end_color=color_light_blue, fill_type="solid")
color_dark_blue = "394D6D"
fill_dark_blue = openpyxl.styles.PatternFill(start_color=color_dark_blue, end_color=color_dark_blue, fill_type="solid")
color_white = "FFFFFF"
fill_white = openpyxl.styles.PatternFill(start_color=color_white, end_color=color_white, fill_type="solid")
color_black = "000000"
fill_black = openpyxl.styles.PatternFill(start_color=color_black, end_color=color_black, fill_type="solid")
color_grey = "D3D3D3"
fill_grey = openpyxl.styles.PatternFill(start_color=color_grey, end_color=color_grey, fill_type="solid")

italic = openpyxl.styles.Font(italic=True)
bold = openpyxl.styles.Font(bold=True)
border = openpyxl.styles.Border(left=openpyxl.styles.Side(border_style='thin', color='000000'))
white_font = openpyxl.styles.Font(color="FFFFFF")

all_base_ids, cobra_model_aerobic, concentration_values_free, concentration_values_paper,\
standardconc_dG0_values, paperconc_dG0_values,\
num_nad_and_nadp_reactions, num_nad_base_ids, num_nadp_base_ids,\
ratio_constraint_data, nad_base_ids, nadp_base_ids, used_growth, zeroed_reaction_ids = load_model_data(anaerobic=False, expanded=False)

all_base_ids, cobra_model_anaerobic, concentration_values_free, concentration_values_paper,\
standardconc_dG0_values, paperconc_dG0_values,\
num_nad_and_nadp_reactions, num_nad_base_ids, num_nadp_base_ids,\
ratio_constraint_data, nad_base_ids, nadp_base_ids, used_growth, zeroed_reaction_ids = load_model_data(anaerobic=True, expanded=False)


wb = openpyxl.Workbook()
current_sheet = 0

########################
# START OF COVER SHEET #
########################

ws = wb.create_sheet("Index")
cell = ws.cell(1, 1)
cell.value = "Supplementary Table 3"
cell.font = bold
cell = ws.cell(2, 1)
cell.value = "of Network-wide Thermodynamic Constraints Shape NAD(P)H Cofactor Specificity of Biochemical Reactions by Pavlos Stephanos Bekiaris¹ & Steffen Klamt¹*"
cell.font = italic
cell = ws.cell(3, 1)
cell.value = "¹Max Planck Institute for Dynamics of Complex Technical Systems, Magdeburg, Sandtorstr. 1, Germany"
cell.font = italic
cell = ws.cell(4, 1)
cell.value = "*Corresponding author: klamt@mpi-magdeburg.mpg.de"
cell.font = italic

cell = ws.cell(6, 1)
cell.value = "INDEX:"
cell.font = bold

cell = ws.cell(7, 1)
cell.value = "Sheet A: We compare the fluxes of a (loopleess) Flux Variability Analysis (FVA) with a thermodynamic FVA under the given growth rate and MDF aerobically and anaerobically, all under standard concentration ranges and with the wild-type specificity"
cell = ws.cell(8, 1)
cell.value = "Sheet B: The same as A, but under the given *Sub*MDF"
cell = ws.cell(9, 1)
cell.value = "Sheet C: We compare the concentration ranges calculated via a Concentration Variability Analysis (CVA) with the measured in vivo data from Bennet et al., 2009 (https://doi.org/10.1038/nchembio.186) under the given growth rate and MDF aerobically, all under standard concentration ranges and with the wild-type specificity"
cell = ws.cell(10, 1)
cell.value = "Sheet D: The same as C, but under the given *Sub*MDF"

cell = ws.cell(12, 1)
cell.value = "All (Sub)MDF are in kJ/mol, all growth rates (µ) in 1/h and all concentrations in M"

########################
# END OF COVER SHEET   #
########################

sheet_to_letter = {
    0: "A",
    1: "B",
    2: "C",
    3: "D",
    4: "E",
    5: "F",
    6: "G",
    7: "H",
    8: "I",
}

for concentrations in ("STANDARDCONCS",): #"PAPERCONCS"):
    original_cobra_model_aerobic = copy.deepcopy(cobra_model_aerobic)
    original_cobra_model_anaerobic = copy.deepcopy(cobra_model_anaerobic)

    cobra_model_aerobic = copy.deepcopy(original_cobra_model_aerobic)
    cobra_model_anaerobic = copy.deepcopy(original_cobra_model_anaerobic)

    condition_models = {
        "Wild-type, aerobic, µ=0.818": cosa_get_model_with_nadx_scenario(nadx_scenario="WILDTYPE", cobra_model=cobra_model_aerobic),
        "Wild-type, anaerobic, µ=0.321": cosa_get_model_with_nadx_scenario(nadx_scenario="WILDTYPE", cobra_model=cobra_model_anaerobic),
    }

    json_tuples = [
        (
            "Wild-type, aerobic, µ=0.818",
            json_load(f"cosa/variability__aerobic_TEST_0_818_{concentrations}_WILDTYPE.json"),
            json_load(f"cosa/variability_STOICHIOMETRIC__aerobic_TEST_0_818_WILDTYPE.json"),
        ),
        (
            "Wild-type, anaerobic, µ=0.321",
            json_load(f"cosa/variability__anaerobic_TEST_0_321_{concentrations}_WILDTYPE.json"),
            json_load(f"cosa/variability_STOICHIOMETRIC__anaerobic_TEST_0_321_WILDTYPE.json"),
        ),
    ]

    json_metadata = {}
    for json_tuple in json_tuples:
        condition = json_tuple[0]
        data = json_tuple[1]
        for target in ("OptMDF", "OptSubMDF"):
            if target not in json_metadata.keys():
                json_metadata[target] = {}
            for elements in data[target]:
                var_id = elements[0]
                min_value = elements[1]
                max_value = elements[2]
                if var_id not in json_metadata[target].keys():
                    json_metadata[target][var_id] = {}

                json_metadata[target][var_id][condition] = {}
                json_metadata[target][var_id][condition]["min"] = min_value
                json_metadata[target][var_id][condition]["max"] = max_value
        data_stoich = json_tuple[2]
        for target in ("OptMDF", "OptSubMDF"):
            for elements in data_stoich["stoichiometric"]:
                var_id = elements[0]
                min_value = elements[1]
                max_value = elements[2]

                json_metadata[target][var_id][condition]["min_stoich"] = min_value
                json_metadata[target][var_id][condition]["max_stoich"] = max_value

    for target in ("OptMDF", "OptSubMDF"):
        ##########################
        ### START OF FVA SHEET ###
        ##########################
        concentrations_string = "Measured" if "PAPER" in concentrations else "Standard"
        ws = wb.create_sheet(sheet_to_letter[current_sheet]+"_"+target.replace("Opt", "")+"_FVAs_"+concentrations_string)
        current_sheet += 1

        cell = ws.cell(1, 1)
        cell.value = f"Comparison of flux ranges from (loopless) Flux Variability Analysis (FVA) and Thermodynamic Flux Variability Analysis at given {target} (TFVA), and maximal driving force for each central carbon metabolism and NAD(P)(H)-dependent reaction, all under standard concentration ranges"

        targetdata = json_metadata[target]
        varnames = list(targetdata.keys())
        f_vars = [x for x in varnames if x.startswith("f_var_")]
        f_vars_core = []
        for f_var in f_vars:
            reac_var = f_var.replace("f_var_", "")
            if "_VARIANT_" in reac_var:
                continue
            reac_var_cleaned = reac_var.replace("_FWD", "").replace("_REV", "").replace("_ORIGINAL_NAD", "").replace("_ORIGINAL_NADP", "").replace("_VARIANT_NAD", "").replace("_VARIANT_NADP", "").replace("_TCOSA", "")
            found = False
            for core_id in core_ids:
                if reac_var_cleaned == core_id:
                    f_vars_core.append(f_var)
                    found = True
                    break
        f_vars_noncore = [x for x in f_vars if (x not in f_vars_core) and ("_TCOSA" in x) and ("_ORIGINAL" in x)]
        f_vars_core.sort()
        f_vars_noncore.sort()
        f_vars = f_vars_core + f_vars_noncore


        cell = ws.cell(4, 1)
        cell.value = "Reaction ID"
        cell.font = italic

        cell = ws.cell(4, 2)
        cell.value = "Reaction string"
        cell.font = italic

        tested_MDFs = {
            condition: round(targetdata["var_B" if target=="OptMDF" else "var_B2"][condition]["max"], 4)
            for condition in targetdata["var_B"].keys()
        }

        headers = list(targetdata[varnames[0]].keys())
        headers_to_column = {}
        n_column = 3
        for header in headers:
            headers_to_column[header] = n_column
            cell = ws.cell(2, n_column)
            cell.value = header
            cell.font = bold
            cell.border = border
            cell = ws.cell(4, n_column)
            cell.value = "min FVA"
            cell.border = border
            cell.font = italic
            cell = ws.cell(4, n_column+1)
            cell.value = "max FVA"
            cell.font = italic
            cell = ws.cell(4, n_column+2)
            cell.value = "min TFVA"
            cell.font = italic
            cell = ws.cell(4, n_column+3)
            cell.value = "max TFVA"
            cell.font = italic
            cell = ws.cell(4, n_column+4)
            cell.value = "max df"
            cell.font = italic

            cell = ws.cell(3, n_column)
            cell.value = "MDF:"
            if target == "OptMDF":
                cell.font = bold
            cell.border = border
            cell = ws.cell(3, n_column+1)
            cell.value = targetdata["var_B"][header]["max"]

            cell = ws.cell(3, n_column+2)
            cell.value = "SubMDF:"
            if target == "OptSubMDF":
                cell.font = bold
            cell = ws.cell(3, n_column+3)
            cell.value = targetdata["var_B2"][header]["max"]

            if ("STANDARD" in concentrations) and (target == "OptMDF") and ("aerobic" in json_tuple[0]):
                standard_optmdf_mdf = targetdata["var_B"][header]["max"]
                standard_optmdf_submdf = targetdata["var_B2"][header]["max"]

            if ("STANDARD" in concentrations) and (target == "OptSubMDF") and ("aerobic" in json_tuple[0]):
                standard_optsubmdf_mdf = targetdata["var_B"][header]["max"]
                standard_optsubmdf_submdf = targetdata["var_B2"][header]["max"]

            mdf = tested_MDFs[header]

            n_column += 5

        max_width = 5
        max_width_str = 5
        current_line = 5

        cell = ws.cell(2, n_column)
        cell.value = "LEGEND:"
        cell.font = bold

        cell = ws.cell(3, n_column)
        cell.value = "Blocked in model"
        cell.fill = fill_black
        cell.font = white_font


        cell = ws.cell(4, n_column)
        cell.value = "Can always run"
        cell.fill = fill_green


        cell = ws.cell(3, n_column+1)
        cell.value = "Blocked in TFVA only"
        cell.fill = fill_light_red


        cell = ws.cell(4, n_column+1)
        cell.value = "Blocked already in FVA"
        cell.fill = fill_dark_red

        cell = ws.cell(3, n_column+2)
        cell.value = "Essential in TFVA only"
        cell.fill = fill_light_blue

        cell = ws.cell(4, n_column+2)
        cell.value = "Essential already in FVA"
        cell.fill = fill_dark_blue


        cell = ws.cell(3, n_column+3)
        cell.value = "Bold text: Bottleneck in at least one condition"
        cell.font = bold

        ws.column_dimensions['C'].width = 10
        ws.column_dimensions['D'].width = 10
        ws.column_dimensions['E'].width = 10
        ws.column_dimensions['F'].width = 10
        ws.column_dimensions['G'].width = 10
        ws.column_dimensions['H'].width = 10
        ws.column_dimensions['I'].width = 10
        ws.column_dimensions['J'].width = 10
        ws.column_dimensions['K'].width = 10
        ws.column_dimensions['L'].width = 10
        ws.column_dimensions['M'].width = 18
        ws.column_dimensions['N'].width = 18
        ws.column_dimensions['O'].width = 23
        ws.column_dimensions['P'].width = 43


        for f_var in f_vars:
            reac_id = f_var.replace("f_var_", "")
            written_reac_id = reac_id.replace("_ORIGINAL_NAD_TCOSA", "")
            written_reac_id = written_reac_id.replace("_ORIGINAL_NADP_TCOSA", "")

            max_width = max(len(written_reac_id)*1.2, max_width)
            ws.column_dimensions['A'].width = max_width

            cell = ws.cell(current_line, 1)
            cell.value = written_reac_id
            cell.alignment = openpyxl.styles.Alignment(horizontal='right')

            cell = ws.cell(current_line, 2)
            cell.value = condition_models[condition].reactions.get_by_id(reac_id).reaction
            max_width_str = max(len(str(condition_models[condition].reactions.get_by_id(reac_id).reaction))*0.5, max_width_str)
            cell.alignment = openpyxl.styles.Alignment(horizontal='right')
            ws.column_dimensions['B'].width = max_width_str

            for condition in targetdata[f_var].keys():
                model_min_flux = condition_models[condition].reactions.get_by_id(reac_id).lower_bound
                model_max_flux = condition_models[condition].reactions.get_by_id(reac_id).upper_bound

                min_fva_flux_value = targetdata[reac_id][condition]["min_stoich"]
                max_fva_flux_value = targetdata[reac_id][condition]["max_stoich"]

                min_tfva_flux_value = targetdata[reac_id][condition]["min"]
                max_tfva_flux_value = targetdata[reac_id][condition]["max"]

                min_df_value = targetdata[f_var][condition]["min"]
                max_df_value = targetdata[f_var][condition]["max"]

                # Is blocked anyway: Black
                if (model_min_flux == 0.0) and (model_max_flux == 0.0):
                    cell_filler = fill_black
                # Is essential in FVA? -> Dark blue
                elif min_fva_flux_value > 1e-8:
                    cell_filler = fill_dark_blue
                # Is essential in TFVA only? -> Light blue
                elif min_tfva_flux_value > 1e-8:
                    cell_filler = fill_light_blue
                # Is blocked?
                elif max_tfva_flux_value < 1e-8:
                    # Blocked in FVA too? -> Dark red
                    if max_fva_flux_value < 1e-8:
                        cell_filler = fill_dark_red
                    # Blocked in TFVA only? -> Light red
                    else:
                        cell_filler = fill_light_red
                # Is active?
                else:
                    cell_filler = fill_green

                cell = ws.cell(current_line, headers_to_column[condition])
                cell.value = round(min_fva_flux_value, 8)
                cell.fill = cell_filler
                cell.border = border
                #
                cell = ws.cell(current_line, headers_to_column[condition]+1)
                cell.value = round(max_fva_flux_value, 8)
                cell.fill = cell_filler

                cell = ws.cell(current_line, headers_to_column[condition]+2)
                cell.value = round(min_tfva_flux_value, 8)
                cell.fill = cell_filler
                #
                cell = ws.cell(current_line, headers_to_column[condition]+3)
                cell.value = round(max_tfva_flux_value, 8)
                cell.fill = cell_filler

                cell = ws.cell(current_line, headers_to_column[condition]+4)
                cell.value = round(max_df_value, 8)
                cell.fill = cell_filler

                if (round(max_df_value, 4) == tested_MDFs[condition]):
                    cell.font = bold

                    cell = ws.cell(current_line, 1)
                    cell.font = bold

            current_line += 1
        ws.freeze_panes = "A5"

        ##########################
        ### END OF FVA SHEET   ###
        ##########################

invivodata = json_load("resources/in_vivo_concentration_data/final_concentration_values_paper.json")
invivodata["nad_tcosa_c"] =   { "min": 2.32e-3, "max":  2.8e-3 }
invivodata["nadh_tcosa_c"] =  { "min": 5.45e-5, "max": 1.27e-4 }
invivodata["nadp_tcosa_c"] =  { "min": 1.4e-7, "max": 3.11e-5 }
invivodata["nadph_tcosa_c"] = { "min": 1.1e-4, "max": 1.34e-4 }

invivo_met_ids = list(invivodata.keys())
invivo_min_concs = {
    met_id: invivodata[met_id]["min"]
    for met_id in invivo_met_ids
}
invivo_max_concs = {
    met_id: invivodata[met_id]["max"]
    for met_id in invivo_met_ids
}

for target in ("OptMDF", "OptSubMDF"):
    ############################
    ### START OF CVA SHEET   ###
    ############################
    ws = wb.create_sheet(sheet_to_letter[current_sheet]+"_"+target.replace("Opt", "")+"_CVAs_STANDARD")
    current_sheet += 1

    cvadata = json_load(f"cosa/results_aerobic/cva_{target.upper()}_STANDARDCONC.json")
    cva_met_ids = list(cvadata.keys())
    cva_met_ids.sort()
    tabledata = {
        met_id: cvadata[met_id]["0,818"]
        for met_id in cva_met_ids
    }
    cva_min_concs = {
        met_id: tabledata[met_id]["min"]
        for met_id in cva_met_ids
    }
    cva_max_concs = {
        met_id: tabledata[met_id]["max"]
        for met_id in cva_met_ids
    }

    ws.freeze_panes = "A4"

    if target == "OptMDF":
        written_mdf = standard_optmdf_mdf
        written_submdf = standard_optmdf_submdf
    else:
        written_mdf = standard_optsubmdf_mdf
        written_submdf = standard_optsubmdf_submdf

    cell = ws.cell(1, 1)
    cell.value = f"Comparison of aerobic concentration ranges from Concentration Variability Analysis at given {target} (CVA) under standard concentration ranges and measured in vivo data from Bennet et al. (2009), https://doi.org/10.1038/nchembio.186"

    cell = ws.cell(2, 1)
    cell.value = "MDF:"
    cell = ws.cell(2, 2)
    cell.value = written_mdf
    cell = ws.cell(2, 3)
    cell.value = "SubMDF:"
    cell = ws.cell(2, 4)
    cell.value = written_submdf
    cell = ws.cell(2, 5)
    cell.value = "µ:"
    cell = ws.cell(2, 6)
    cell.value = 0.818

    cell = ws.cell(3, 1)
    cell.value = "Metabolite ID"
    cell.font = italic
    cell = ws.cell(3, 2)
    cell.value = "Metabolite Name"
    cell.font = italic
    cell = ws.cell(3, 3)
    cell.value = "Min CVA conc."
    cell.font = italic
    cell = ws.cell(3, 4)
    cell.value = "Max CVA conc."
    cell.font = italic
    cell = ws.cell(3, 5)
    cell.value = "Min in vivo conc."
    cell.font = italic
    cell = ws.cell(3, 6)
    cell.value = "Max in vivo conc."
    cell.font = italic

    cell = ws.cell(2, 7)
    cell.value = "LEGEND:"
    cell.font = bold

    cell = ws.cell(3, 7)
    cell.value = "CVA range not in measured range"
    cell.fill = fill_light_red

    cell = ws.cell(3, 8)
    cell.value = "CVA range in measured range"
    cell.fill = fill_green

    # cell = ws.cell(3, 9)
    # cell.value = "CVA data only"
    # cell.fill = fill_grey

    current_line = 4
    for cva_met_id in cva_met_ids:
        cva_min_conc = cva_min_concs[cva_met_id]
        cva_max_conc = cva_max_concs[cva_met_id]

        converted_met_id = cva_met_id[2:]  # No "x_" var prefix

        if converted_met_id in invivo_met_ids:
            invivo_min_conc = invivo_min_concs[converted_met_id]
            invivo_max_conc = invivo_max_concs[converted_met_id]

            if ((cva_min_conc >= invivo_min_conc) and (cva_max_conc <= invivo_max_conc)) or ((cva_min_conc <= invivo_min_conc) and (cva_max_conc >= invivo_max_conc)) or ((cva_min_conc <= invivo_min_conc) and (cva_max_conc >= invivo_min_conc)) or ((invivo_min_conc <= invivo_max_conc) and (cva_max_conc >= invivo_max_conc)):
                cell_filler = fill_green
            else:
                cell_filler = fill_light_red
        elif converted_met_id == "nad_tcosa_c":
            invivo_min_conc = "N/A"
            invivo_max_conc = "N/A"
            cell_filler = fill_grey

        cell = ws.cell(current_line, 1)
        cell.value = converted_met_id
        cell.fill = cell_filler

        cell = ws.cell(current_line, 2)
        cell.value = cobra_model_aerobic.metabolites.get_by_id(converted_met_id).name
        cell.fill = cell_filler

        cell = ws.cell(current_line, 3)
        cell.value = cva_min_conc
        cell.fill = cell_filler

        cell = ws.cell(current_line, 4)
        cell.value = cva_max_conc
        cell.fill = cell_filler

        cell = ws.cell(current_line, 5)
        cell.value = invivo_min_conc
        cell.fill = cell_filler

        cell = ws.cell(current_line, 6)
        cell.value = invivo_max_conc
        cell.fill = cell_filler

        current_line += 1

    ws.column_dimensions['A'].width = 18
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 16
    ws.column_dimensions['D'].width = 16
    ws.column_dimensions['E'].width = 16
    ws.column_dimensions['F'].width = 16
    ws.column_dimensions['G'].width = 25
    ws.column_dimensions['H'].width = 25
    ws.column_dimensions['I'].width = 15


del(wb["Sheet"])
wb.save("./cosa/Supplementary_Table_3.xlsx")
